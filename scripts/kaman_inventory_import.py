#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Import inventory + links into Kaman Manager API (thex.kaman.dev).

Data sources (public/):
  - 1774464381.csv              → inventory items + receive-stock
  - סיכום חומרי גלם (1).xlsx    → ingredient names + default qty hints
  - ניתוח פוד קוסט ...xlsx      → dish ↔ component qty (link-item / link-ingredient)

Usage:
  python scripts/kaman_inventory_import.py --dry-run
  python scripts/kaman_inventory_import.py --execute --limit 25
  python scripts/kaman_inventory_import.py --execute --suppliers-map scripts/suppliers_map.example.json

Requires: openpyxl (pip install openpyxl)
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import ssl
import sys
import time
import unicodedata
import urllib.error
import urllib.request
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
PUBLIC = ROOT / "public"
OUT_DIR = Path(__file__).resolve().parent / "output"

DEFAULT_BASE = "https://thex.kaman.dev/api/manager"
DEFAULT_EMAIL = "thex@kaman.rest"
DEFAULT_PASSWORD = "1234"

CSV_PATH = PUBLIC / "1774464381.csv"
RAW_XLSX = PUBLIC / "סיכום חומרי גלם (1).xlsx"
FOOD_XLSX = PUBLIC / "ניתוח פוד קוסט תאורתי ברטעה ינואר 26 (1).xlsx"


def norm_name(value: str) -> str:
    if not value:
        return ""
    s = unicodedata.normalize("NFKC", str(value)).strip().lower()
    s = s.replace("״", '"').replace("׳", "'").replace("”", '"').replace("“", '"')
    s = re.sub(r"\s+", " ", s)
    return s


def map_unit(hebrew: str) -> str:
    u = norm_name(hebrew)
    mapping = {
        'ק"ג': "kg",
        "קג": "kg",
        "kg": "kg",
        "גרם": "g",
        "g": "g",
        "ליטר": "l",
        "l": "l",
        'מ"ל': "ml",
        "מל": "ml",
        "ml": "ml",
        "יחידות": "piece",
        "unit": "piece",
        "units": "piece",
        "גלון": "piece",
        "קרטון": "box",
        "שקית": "bag",
        'סמ"ק': "piece",
        "נספק": "piece",
    }
    return mapping.get(u, "piece")


def clean_sku(raw: str) -> str:
    s = re.sub(r"[^\w\-]+", "", str(raw or "").strip(), flags=re.UNICODE)
    return s[:64] if s else ""


@dataclass
class CsvRow:
    name: str
    sku: str
    pack_qty: float
    unit: str
    supplier: str
    price: float


def load_csv_rows() -> list[CsvRow]:
    rows: list[CsvRow] = []
    with open(CSV_PATH, encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f):
            name = (r.get("שם מוצר") or "").strip()
            if not name:
                continue
            try:
                qty = float(r.get("כמות") or 1)
            except ValueError:
                qty = 1.0
            try:
                price = float(r.get("מחיר מחירון") or 0)
            except ValueError:
                price = 0.0
            sku = clean_sku(r.get('מק"ט') or "")
            rows.append(
                CsvRow(
                    name=name,
                    sku=sku or clean_sku(name)[:32],
                    pack_qty=max(qty, 0.001),
                    unit=map_unit(r.get("יחידת מידה") or ""),
                    supplier=(r.get("ספק") or "").strip(),
                    price=max(price, 0),
                )
            )
    return rows


def load_raw_material_names() -> list[tuple[str, float]]:
    out: list[tuple[str, float]] = []
    wb = openpyxl.load_workbook(RAW_XLSX, data_only=True, read_only=True)
    sh = wb[wb.sheetnames[0]]
    for i, row in enumerate(sh.iter_rows(values_only=True)):
        if i < 2:
            continue
        name = row[1] if len(row) > 1 else None
        qty = row[2] if len(row) > 2 else None
        if not name or not str(name).strip():
            continue
        try:
            q = float(qty) if qty is not None else 0.0
        except (TypeError, ValueError):
            q = 0.0
        out.append((str(name).strip(), q))
    wb.close()
    return out


def load_food_cost_links() -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(FOOD_XLSX, data_only=True, read_only=True)
    links: list[dict[str, Any]] = []
    for sheet_name in wb.sheetnames:
        if sheet_name == "📊 Summary":
            continue
        sh = wb[sheet_name]
        current_dish = None
        for i, row in enumerate(sh.iter_rows(values_only=True), start=1):
            if i < 3:
                continue
            cells = list(row)
            if not any(c is not None and str(c).strip() for c in cells[:6]):
                continue
            dish = cells[0]
            component = cells[2] if len(cells) > 2 else None
            typ = cells[3] if len(cells) > 3 else None
            qty = cells[4] if len(cells) > 4 else None
            if dish and str(dish).strip():
                current_dish = str(dish).strip()
            if not current_dish or not component or not str(component).strip():
                continue
            try:
                q = float(qty)
            except (TypeError, ValueError):
                continue
            links.append(
                {
                    "dish": current_dish,
                    "component": str(component).strip(),
                    "type": str(typ or "").strip(),
                    "qty": q,
                }
            )
    wb.close()
    return links


class KamanClient:
    def __init__(self, base_url: str, email: str, password: str, verify_ssl: bool = False):
        self.base_url = base_url.rstrip("/")
        self.ctx = ssl.create_default_context()
        if not verify_ssl:
            self.ctx.check_hostname = False
            self.ctx.verify_mode = ssl.CERT_NONE
        self.token = self._login(email, password)
        self.inventory_by_name: dict[str, str] = {}
        self.inventory_by_sku: dict[str, str] = {}
        self.ingredient_by_name: dict[str, str] = {}
        self.item_by_name: dict[str, str] = {}
        self.ingredient_category_id: str | None = None

    def _login(self, email: str, password: str) -> str:
        data = json.dumps({"email": email, "password": password}).encode()
        req = urllib.request.Request(
            f"{self.base_url}/login",
            data=data,
            headers={"Accept": "application/json", "Content-Type": "application/json"},
            method="POST",
        )
        with urllib.request.urlopen(req, context=self.ctx) as resp:
            body = json.loads(resp.read().decode())
        token = body.get("data", {}).get("token") or body.get("token")
        if not token:
            raise RuntimeError(f"Login failed: {body}")
        return token

    def request(self, method: str, path: str, payload: dict | None = None) -> tuple[int, dict]:
        headers = {
            "Accept": "application/json",
            "Authorization": f"Bearer {self.token}",
        }
        data = None
        if payload is not None:
            headers["Content-Type"] = "application/json"
            data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        req = urllib.request.Request(
            f"{self.base_url}/{path.lstrip('/')}",
            data=data,
            headers=headers,
            method=method,
        )
        try:
            with urllib.request.urlopen(req, context=self.ctx) as resp:
                raw = resp.read().decode()
                return resp.status, json.loads(raw) if raw else {}
        except urllib.error.HTTPError as e:
            raw = e.read().decode()
            try:
                return e.code, json.loads(raw)
            except json.JSONDecodeError:
                return e.code, {"message": raw}

    def ensure_ingredient_category(self) -> str:
        if self.ingredient_category_id:
            return self.ingredient_category_id
        existing = self.request("GET", "ingredients-categories")
        cats = existing[1].get("data") or []
        for c in cats:
            name = norm_name(c.get("name_he") or c.get("name_en") or c.get("name_ar") or "")
            if name in (norm_name("חומרי גלם"), norm_name("materials"), norm_name("مواد")):
                self.ingredient_category_id = c["id"]
                return self.ingredient_category_id
        code, body = self.request(
            "POST",
            "ingredients-categories",
            {"name_ar": "مواد خام", "name_en": "Raw materials", "name_he": "חומרי גלם"},
        )
        if code >= 400:
            raise RuntimeError(f"Create ingredients category failed: {body}")
        self.ingredient_category_id = body["data"]["id"]
        return self.ingredient_category_id

    def refresh_items(self) -> None:
        code, body = self.request("GET", "items")
        if code >= 400:
            return
        for item in body.get("data") or []:
            for key in ("name_en", "name_he", "name_ar"):
                v = item.get(key)
                if v:
                    self.item_by_name[norm_name(v)] = item["id"]

    def refresh_inventory(self) -> None:
        code, body = self.request("GET", "inventory-items")
        if code >= 400:
            return
        for inv in body.get("data") or []:
            inv_id = inv.get("id")
            if not inv_id:
                continue
            sku = inv.get("sku")
            if sku:
                sku_key = clean_sku(str(sku))
                if sku_key:
                    self.inventory_by_sku[sku_key] = inv_id
            for key in ("name_en", "name_he", "name_ar"):
                v = inv.get(key)
                if v:
                    self.inventory_by_name[norm_name(v)] = inv_id

    def refresh_ingredients(self) -> None:
        code, body = self.request("GET", "ingredients")
        if code >= 400:
            return
        for ing in body.get("data") or []:
            for key in ("name_en", "name_he", "name_ar"):
                v = ing.get(key)
                if v:
                    self.ingredient_by_name[norm_name(v)] = ing["id"]

    def create_inventory(self, row: CsvRow, sku_suffix: str = "") -> str | None:
        key = norm_name(row.name)
        if key in self.inventory_by_name:
            return self.inventory_by_name[key]
        sku = (row.sku + sku_suffix)[:64]
        sku_key = clean_sku(sku)
        if sku_key and sku_key in self.inventory_by_sku:
            inv_id = self.inventory_by_sku[sku_key]
            self.inventory_by_name[key] = inv_id
            return inv_id
        payload = {
            "name_ar": row.name,
            "name_en": row.name,
            "name_he": row.name,
            "sku": sku,
            "unit": row.unit,
            "minimum_quantity": 1,
            "track_inventory": True,
            "depreciation_rate": 0,
        }
        code, body = self.request("POST", "inventory-items", payload)
        if code == 422:
            self.refresh_inventory()
            if key in self.inventory_by_name:
                return self.inventory_by_name[key]
            if sku_key and sku_key in self.inventory_by_sku:
                inv_id = self.inventory_by_sku[sku_key]
                self.inventory_by_name[key] = inv_id
                return inv_id
            err = json.dumps(body, ensure_ascii=False)
            if sku_suffix == "" and "sku" in err.lower():
                return self.create_inventory(row, sku_suffix="-" + sku[-6:])
        if code == 429:
            time.sleep(2)
            return self.create_inventory(row, sku_suffix)
        if code >= 400:
            print(f"  [inventory FAIL] {row.name}: {body.get('message') or body}", file=sys.stderr)
            return None
        inv_id = body["data"]["id"]
        self.inventory_by_name[key] = inv_id
        return inv_id

    def create_ingredient(self, name: str, price: float = 0) -> str | None:
        key = norm_name(name)
        if key in self.ingredient_by_name:
            return self.ingredient_by_name[key]
        cat_id = self.ensure_ingredient_category()
        payload = {
            "name_ar": name,
            "name_en": name,
            "name_he": name,
            "price": price,
            "category_id": cat_id,
        }
        code, body = self.request("POST", "ingredients", payload)
        if code == 429:
            time.sleep(2)
            return self.create_ingredient(name, price)
        if code >= 400:
            print(f"  [ingredient FAIL] {name}: {body.get('message') or body}", file=sys.stderr)
            return None
        ing_id = body["data"]["id"]
        self.ingredient_by_name[key] = ing_id
        return ing_id

    def link_inventory_ingredient(self, inv_id: str, ing_id: str, qty: float) -> bool:
        code, body = self.request(
            "POST",
            f"inventory-items/{inv_id}/link-ingredient",
            {"ingredient_id": ing_id, "quantity_consumed": round(qty, 6)},
        )
        return code < 400

    def link_inventory_item(self, inv_id: str, item_id: str, qty: float) -> bool:
        code, body = self.request(
            "POST",
            f"inventory-items/{inv_id}/link-item",
            {"item_id": item_id, "quantity_consumed": round(qty, 6)},
        )
        return code < 400

    def receive_stock(
        self, inv_id: str, supplier_id: str, quantity: float, price_per_unit: float, notes: str = ""
    ) -> bool:
        payload = {
            "supplier_id": supplier_id,
            "quantity": round(quantity, 4),
            "price_per_unit": round(price_per_unit, 4),
            "notes": notes[:250],
        }
        code, body = self.request("POST", f"inventory-items/{inv_id}/receive-stock", payload)
        if code >= 400:
            print(f"  [stock FAIL] {inv_id}: {body.get('message') or body}", file=sys.stderr)
        return code < 400


def write_http_file(path: Path, base_url: str, token_placeholder: str = "{{token}}") -> None:
    """Write a REST Client .http file with the main request templates."""
    content = f"""# Kaman inventory import — thex (dev)
# Set token after login:
# POST {base_url}/login
@baseUrl = {base_url}
@token = {token_placeholder}
@inventoryItemId = 
@ingredientId = 
@itemId = 
@supplierId = 

### Login
POST {{{{baseUrl}}}}/login
Content-Type: application/json
Accept: application/json

{{
  "email": "thex@kaman.rest",
  "password": "1234"
}}

### 0) List inventory items
GET {{{{baseUrl}}}}/inventory-items
Authorization: Bearer {{{{token}}}}

### 1) Create inventory item
POST {{{{baseUrl}}}}/inventory-items
Authorization: Bearer {{{{token}}}}
Content-Type: application/json
Accept: application/json

{{
  "name_ar": "דוגמה",
  "name_en": "Example",
  "name_he": "דוגמה",
  "sku": "SKU-001",
  "unit": "kg",
  "minimum_quantity": 1,
  "track_inventory": true,
  "depreciation_rate": 0
}}

### 2) Link inventory → ingredient
POST {{{{baseUrl}}}}/inventory-items/{{{{inventoryItemId}}}}/link-ingredient
Authorization: Bearer {{{{token}}}}
Content-Type: application/json

{{
  "ingredient_id": "{{{{ingredientId}}}}",
  "quantity_consumed": 0.15
}}

### 3) Link inventory → menu item
POST {{{{baseUrl}}}}/inventory-items/{{{{inventoryItemId}}}}/link-item
Authorization: Bearer {{{{token}}}}
Content-Type: application/json

{{
  "item_id": "{{{{itemId}}}}",
  "quantity_consumed": 0.2
}}

### 4) Add stock quantity (needs valid supplier_id from Kaman UI)
POST {{{{baseUrl}}}}/inventory-items/{{{{inventoryItemId}}}}/receive-stock
Authorization: Bearer {{{{token}}}}
Content-Type: application/json

{{
  "supplier_id": "{{{{supplierId}}}}",
  "quantity": 10,
  "price_per_unit": 16.99,
  "received_date": "2026-05-24",
  "notes": "Import from CSV"
}}
"""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Import Kaman inventory from uploaded files")
    parser.add_argument("--base-url", default=DEFAULT_BASE)
    parser.add_argument("--email", default=DEFAULT_EMAIL)
    parser.add_argument("--password", default=DEFAULT_PASSWORD)
    parser.add_argument("--dry-run", action="store_true", help="Only write HTTP template + plan counts")
    parser.add_argument("--execute", action="store_true", help="Call the API")
    parser.add_argument("--limit", type=int, default=0, help="Max CSV rows to import (0 = all)")
    parser.add_argument("--skip-stock", action="store_true", help="Skip receive-stock calls")
    parser.add_argument(
        "--suppliers-map",
        type=Path,
        help="JSON map: supplier display name -> supplier UUID (create suppliers in Kaman UI first)",
    )
    parser.add_argument("--skip-links", action="store_true", help="Only create inventory + ingredients")
    args = parser.parse_args()

    csv_rows = load_csv_rows()
    raw_materials = load_raw_material_names()
    recipe_links = load_food_cost_links()

    print(f"CSV inventory rows: {len(csv_rows)}")
    print(f"Raw material sheet rows: {len(raw_materials)}")
    print(f"Food-cost component links: {len(recipe_links)}")

    http_path = OUT_DIR / "kaman_inventory_import.http"
    write_http_file(http_path, args.base_url)
    print(f"Wrote HTTP templates: {http_path}")

    if args.dry_run and not args.execute:
        print("Dry run only. Use --execute to call the API (optionally with --limit).")
        return 0

    if not args.execute:
        print("Nothing executed. Pass --execute to import.")
        return 0

    supplier_map: dict[str, str] = {}
    if args.suppliers_map and args.suppliers_map.exists():
        supplier_map = {
            norm_name(k): v for k, v in json.loads(args.suppliers_map.read_text(encoding="utf-8")).items()
        }
    elif not args.skip_stock:
        print(
            "Warning: no --suppliers-map; receive-stock will be skipped. "
            "Create suppliers in Kaman UI and provide name→UUID JSON.",
            file=sys.stderr,
        )
        args.skip_stock = True

    client = KamanClient(args.base_url, args.email, args.password)
    client.refresh_inventory()
    client.refresh_items()
    client.refresh_ingredients()

    limit = args.limit if args.limit > 0 else len(csv_rows)
    created_inv = 0
    stocked = 0

    print(f"\n=== Phase 1: inventory items (limit {limit}) ===")
    for row in csv_rows[:limit]:
        inv_id = client.create_inventory(row)
        if not inv_id:
            continue
        created_inv += 1
        if not args.skip_stock and row.supplier:
            sid = supplier_map.get(norm_name(row.supplier))
            if sid:
                if client.receive_stock(inv_id, sid, row.pack_qty, row.price, notes=row.supplier):
                    stocked += 1
            else:
                print(f"  [skip stock] unknown supplier: {row.supplier}", file=sys.stderr)
        time.sleep(0.2)

    print(f"Created/resolved inventory: {created_inv}, receive-stock OK: {stocked}")

    print("\n=== Phase 2: ingredients from raw-materials summary ===")
    ing_created = 0
    for name, _qty in raw_materials:
        if client.create_ingredient(name):
            ing_created += 1
        time.sleep(0.2)
    print(f"Ingredients created/resolved: {ing_created}")

    print("\n=== Phase 3: link inventory -> ingredient (name match) ===")
    linked_ing = 0
    for name, _qty in raw_materials:
        inv_id = client.inventory_by_name.get(norm_name(name))
        ing_id = client.ingredient_by_name.get(norm_name(name))
        if inv_id and ing_id:
            qty = _qty if _qty > 0 else 0.01
            if client.link_inventory_ingredient(inv_id, ing_id, qty):
                linked_ing += 1
    print(f"inventory<->ingredient links: {linked_ing}")

    if args.skip_links:
        return 0

    print("\n=== Phase 4: food-cost links (inventory -> menu item) ===")
    linked_items = 0
    missing_dish = 0
    missing_component = 0
    for link in recipe_links:
        dish_key = norm_name(link["dish"])
        comp_key = norm_name(link["component"])
        item_id = client.item_by_name.get(dish_key)
        inv_id = client.inventory_by_name.get(comp_key)
        if not item_id:
            missing_dish += 1
            continue
        if not inv_id:
            # try ingredient path: create ingredient + link inventory
            ing_id = client.ingredient_by_name.get(comp_key) or client.create_ingredient(link["component"])
            inv_id = client.inventory_by_name.get(comp_key)
            if ing_id and inv_id and client.link_inventory_ingredient(inv_id, ing_id, link["qty"]):
                linked_ing += 1
            else:
                missing_component += 1
            continue
        if client.link_inventory_item(inv_id, item_id, link["qty"]):
            linked_items += 1
        time.sleep(0.15)

    print(f"inventory->item links: {linked_items}")
    print(f"(unmatched dishes: {missing_dish}, unmatched components: {missing_component})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
