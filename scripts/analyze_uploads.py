# -*- coding: utf-8 -*-
import csv
import json
import os
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1] / "public"

files = {
    "csv": ROOT / "1774464381.csv",
    "raw_materials": ROOT / "סיכום חומרי גלם (1).xlsx",
    "food_cost": ROOT / "ניתוח פוד קוסט תאורתי ברטעה ינואר 26 (1).xlsx",
}

out = {}

# CSV
with open(files["csv"], encoding="utf-8-sig", newline="") as f:
    reader = csv.DictReader(f)
    rows = list(reader)
out["csv"] = {
    "rows": len(rows),
    "columns": reader.fieldnames,
    "sample": rows[:3],
    "suppliers_unique": sorted({r.get("ספק", "").strip() for r in rows if r.get("ספק")})[:20],
    "categories_unique_count": len({r.get("קטגוריה", "").strip() for r in rows if r.get("קטגוריה")}),
}

# Excel files
for key, path in [("raw_materials", files["raw_materials"]), ("food_cost", files["food_cost"])]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheets = {}
    for name in wb.sheetnames:
        sh = wb[name]
        rows_data = []
        for i, row in enumerate(sh.iter_rows(values_only=True)):
            rows_data.append([("" if c is None else str(c))[:60] for c in row[:20]])
            if i >= 12:
                break
        sheets[name] = {
            "preview_rows": rows_data,
            "max_row": sh.max_row,
            "max_col": sh.max_column,
        }
    wb.close()
    out[key] = {"sheets": sheets, "sheet_names": list(sheets.keys())}

out_path = Path(__file__).resolve().parent / "upload_analysis.json"
out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"Wrote {out_path}")
