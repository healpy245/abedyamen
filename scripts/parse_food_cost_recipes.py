# -*- coding: utf-8 -*-
"""Extract dish -> component qty mappings from food cost workbook."""
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1] / "public"
wb = openpyxl.load_workbook(
    ROOT / "ניתוח פוד קוסט תאורתי ברטעה ינואר 26 (1).xlsx",
    data_only=True,
    read_only=True,
)

recipes = []
skip = {"📊 Summary"}

for sheet_name in wb.sheetnames:
    if sheet_name in skip:
        continue
    sh = wb[sheet_name]
    current_dish = None
    for i, row in enumerate(sh.iter_rows(values_only=True), start=1):
        if i < 3:
            continue
        cells = [c for c in row]
        if not any(c is not None and str(c).strip() != "" for c in cells[:6]):
            continue
        dish = cells[0]
        component = cells[2] if len(cells) > 2 else None
        typ = cells[3] if len(cells) > 3 else None
        qty = cells[4] if len(cells) > 4 else None
        unit = cells[5] if len(cells) > 5 else None

        if dish and str(dish).strip():
            current_dish = str(dish).strip()
        if not current_dish or not component or str(component).strip() == "":
            continue
        if str(typ or "").strip().lower() not in ("item", "recipe", "ingredient", "temporar", "temporary"):
            # still accept Item/Recipe case-insensitive
            t = str(typ or "").strip().lower()
            if t and t not in ("item", "recipe"):
                continue
        try:
            q = float(qty) if qty is not None and str(qty).strip() != "" else None
        except (TypeError, ValueError):
            q = None
        if q is None:
            continue
        recipes.append(
            {
                "sheet": sheet_name,
                "dish": current_dish,
                "component": str(component).strip(),
                "type": str(typ or "").strip(),
                "qty": q,
                "unit": str(unit or "").strip(),
            }
        )

wb.close()

out = {
    "recipe_links": len(recipes),
    "unique_dishes": len({r["dish"] for r in recipes}),
    "unique_components": len({r["component"] for r in recipes}),
    "sample": recipes[:15],
}
Path(__file__).parent.joinpath("food_cost_recipes.json").write_text(
    json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8"
)
print(json.dumps({k: out[k] for k in out if k != "sample"}, ensure_ascii=False))
