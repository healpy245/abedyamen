# -*- coding: utf-8 -*-
import csv
import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1] / "public"
out = {}

# CSV
with open(ROOT / "1774464381.csv", encoding="utf-8-sig", newline="") as f:
    reader = csv.DictReader(f)
    rows = list(reader)
out["csv"] = {
    "rows": len(rows),
    "columns": reader.fieldnames,
    "unit_values": sorted({(r.get("יחידת מידה") or "").strip() for r in rows})[:15],
    "supplier_count": len({(r.get("ספק") or "").strip() for r in rows if (r.get("ספק") or "").strip()}),
}

# Raw materials xlsx
wb = openpyxl.load_workbook(ROOT / "סיכום חומרי גלם (1).xlsx", data_only=True, read_only=True)
sheet_name = wb.sheetnames[0]
sh = wb[sheet_name]
rows = []
for i, row in enumerate(sh.iter_rows(values_only=True)):
    if i > 5:
        break
    rows.append([str(c)[:50] if c is not None else "" for c in row[:8]])
wb.close()
out["raw_materials_xlsx"] = {"sheet": sheet_name, "preview": rows}

# Food cost - count recipe rows
wb2 = openpyxl.load_workbook(ROOT / "ניתוח פוד קוסט תאורתי ברטעה ינואר 26 (1).xlsx", data_only=True, read_only=True)
recipe_stats = {}
for name in wb2.sheetnames:
    sh = wb2[name]
    headers = None
    dish_count = 0
    component_rows = 0
    for i, row in enumerate(sh.iter_rows(values_only=True)):
        if i == 1 and name in ("Appetizers", "beside the burger", "Beverages"):
            headers = [str(c).strip() if c else "" for c in row]
        if i > 1 and row and row[1]:
            if row[0] and str(row[0]).strip():
                dish_count += 1
            if len(row) > 2 and row[2]:
                component_rows += 1
    recipe_stats[name] = {"max_row": sh.max_row, "dish_headers_row2": headers, "dish_count_guess": dish_count, "component_rows_guess": component_rows}
wb2.close()
out["food_cost"] = {"sheets": list(recipe_stats.keys()), "stats": recipe_stats}

Path(__file__).parent.joinpath("upload_summary.json").write_text(
    json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8"
)
print("done")
